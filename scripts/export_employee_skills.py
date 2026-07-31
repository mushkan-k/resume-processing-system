"""
Export employee skills to Excel - Quick export for when frontend is down
Creates: employee_skills_export.xlsx with columns: Name, Skillset

Usage:
    python scripts/export_employee_skills.py
"""

import sys
import os
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from mcp_server.db.connection import get_db_connection as get_connection


def export_to_excel():
    """Export employee skills to Excel file"""
    
    print("Connecting to database...")
    with get_connection() as conn:
        cur = conn.cursor(dictionary=True)
        
        # Get all employees with their skills
        print("Fetching employee data...")
        cur.execute("""
            SELECT 
                e.employee_id,
                e.id,
                e.employee_name,
                GROUP_CONCAT(DISTINCT es.skill ORDER BY es.skill SEPARATOR ' | ') as skillset
            FROM employee e
            LEFT JOIN employee_skillset es ON e.id = es.employee_jobdiva_id
            GROUP BY e.employee_id, e.id, e.employee_name
            ORDER BY e.employee_name
        """)
        
        employees = cur.fetchall()
        cur.close()
    
    print(f"Found {len(employees)} employees")
    
    # Create Excel workbook
    print("Creating Excel file...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Skills"
    
    # Headers
    headers = ["Employee ID", "ID", "Name", "Skillset"]
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    
    # Add data
    for emp in employees:
        ws.append([
            emp["employee_id"],
            emp["id"],
            emp["employee_name"],
            emp["skillset"] or ""
        ])
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)  # Cap at 100
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    output_dir = Path(__file__).parent.parent / "data"
    output_dir.mkdir(exist_ok=True)
    output_file = output_dir / "employee_skills_export.xlsx"
    
    wb.save(output_file)
    print(f"\n✓ Excel file created: {output_file}")
    print(f"  Total employees: {len(employees)}")
    print(f"  Columns: {', '.join(headers)}")
    
    return output_file


if __name__ == "__main__":
    try:
        export_to_excel()
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
